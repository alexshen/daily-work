#! /usr/bin/env python3

import argparse
import openpyxl
import functools
import os
import re
from copy import copy


def copy_styles(ws, src_row, dest_row):
    for i in range(ws.min_column, ws.max_column + 1):
        sc = ws.cell(src_row, i)
        dc = ws.cell(dest_row, i)

        dc.font = copy(sc.font)
        dc.border = copy(sc.border)
        dc.fill = copy(sc.fill)
        dc.number_format = copy(sc.number_format)
        dc.protection = copy(sc.protection)
        dc.alignment = copy(sc.alignment)

    ws.row_dimensions[dest_row].height = ws.row_dimensions[src_row].height


class Resident:
    def __init__(self, name='', id='', phone='', comment=''):
        self.name = name
        self.id = id
        self.phone = phone
        self.comment = comment

class Room:
    def __init__(self, addr, tag=''):
        self.addr = addr
        self.tag = tag
        self.residents = []


class DocumentWriter:
    _IDX_COL = 1
    _ROOM_COL = 2
    _NAME_COL = 3
    _ID_COL = 4
    _PHONE_COL = 5
    _COMMENT_COL = 6
    _ROOM_TAG = 7

    def __init__(self, template, path, has_outline=True, has_room_tag=True, separate_room_tag=True):
        self._template = template
        self._has_outline = has_outline
        self._path = path
        self._has_room_tag = has_room_tag
        self._separate_room_tag = separate_room_tag   
        self._unit_addr = None
        self._rooms = {}

    def set_unit_addr(self, unit_addr):
        self._unit_addr = unit_addr

    def add_room(self, addr, tag=''):
        if addr in self._rooms:
            raise ValueError('duplicate room {0}'.format(addr))
        self._rooms[addr] = Room(addr, tag)

    def add_resident(self, addr, resident):
        try:
            room = self._rooms[addr]
        except KeyError:
            room = self._rooms[addr] = Room(addr)
        room.residents.append(resident)

    @property
    def num_residents(self):
        return functools.reduce(lambda total, e: total + len(e.residents), self._rooms.values(), 0)

    @property
    def num_rooms(self):
        return len(self._rooms)

    def save(self, merge=False, first_resident_address_only=True):
        tmpl = openpyxl.load_workbook(self._template)
        ws = tmpl.active
        # update title
        title_cell = ws.cell(1, 1)
        title_cell.value = title_cell.value.format(
            unit_addr=self._unit_addr)
        # update outline
        outline_cell = ws.cell(2, 1)
        outline_cell.value = outline_cell.value.format(
            unit_addr=self._unit_addr, num_rooms=self.num_rooms, num_residents=self.num_residents)

        idx = 1
        first_row = self._has_outline and 4 or 3
        row = first_row
        for room in self._rooms.values():
            # get all the residents, if there's no resident, we need to add a placeholder resident
            # so that the room can be kept
            residents = room.residents if room.residents else [
                Resident(name='', id='', phone='', comment='')]
            if self._has_room_tag:
                if self._separate_room_tag:
                    ws.cell(row, self._ROOM_TAG).value = room.tag
            start_row = row
            # output residents ordered by name in ascending order
            for i, r in enumerate(sorted(residents, key=lambda x: x.name)):
                ws.cell(row, self._IDX_COL).value = idx
                if i == 0 or not first_resident_address_only:
                    ws.cell(row, self._ROOM_COL).value = room.addr
                ws.cell(row, self._NAME_COL).value = r.name
                ws.cell(row, self._ID_COL).value = r.id
                ws.cell(row, self._PHONE_COL).value = r.phone
                comment = r.comment
                if i == 0 and self._has_room_tag and not self._separate_room_tag:
                    comment += (r.comment and ' ' or '') + room.tag
                ws.cell(row, self._COMMENT_COL).value = comment
                copy_styles(ws, first_row, row)
                row += 1
                idx += 1
            if merge and len(residents) > 1:
                ws.merge_cells(start_row=start_row, start_column=self._ROOM_COL,
                               end_row=start_row + len(residents) - 1, end_column=self._ROOM_COL)

        os.makedirs(os.path.dirname(self._path), exist_ok=True)
        tmpl.save(self._path)


class TableReader:
    def __init__(self, ws_table):
        self._ws_table = ws_table
        self._column_indices = {}
        for i, c in enumerate(next(self._ws_table.rows)):
            self._column_indices[c.value] = i + 1

    @property
    def column_names(self):
        return tuple(self._column_indices.keys())

    @property
    def num_rows(self):
        return self._ws_table.max_row - self._ws_table.min_row

    def row(self, row):
        return TableRow(self, row)

    def __iter__(self):
        for i in range(1, self.num_rows + 1):
            yield self.row(i)

    def value(self, row, column_name):
        # first row is the headers
        return self._ws_table.cell(row + 1, self._column_indices[column_name]).value


class TableRow:
    def __init__(self, reader, row):
        self._reader = reader
        self._row = row

    def __getitem__(self, column_name):
        return self._reader.value(self._row, column_name)


class Exporter:
    def __init__(self, template_path, data_xlsx, output_dir, 
                 has_outline=True, has_room_tag=True, separate_room_tag=True, comment_tags=[], exclude_tags=[]):
        self._template_path = template_path
        self._has_outline = has_outline
        self._has_room_tag = has_room_tag
        self._separate_room_tag = separate_room_tag
        self._db_wb = openpyxl.load_workbook(data_xlsx)
        self._output_dir = output_dir
        self._residents = {}
        self._room_tags = {}
        self._comment_tags = [re.compile(pat) for pat in (comment_tags or [])]
        self._exclude_tags = set(exclude_tags or [])

        self._read_residents()
        self._read_room_tags()

    def _read_residents(self):
        for row in TableReader(self._db_wb['在住']):
            comments = []
            if row['社区标识']:
                tags = set(row['社区标识'].split(','))
                if self._exclude_tags.isdisjoint(tags):
                    for comment_tag in self._comment_tags:
                        for tag in tags:
                            for m in comment_tag.finditer(tag):
                                if len(m.groups()) > 0:
                                    comments.append(m[1])
                                else:
                                    comments.append(m[0])
            r = Resident(row['姓名'], row['身份证'], row['电话'], ' '.join(comments))
            self._residents.setdefault(row['关联房屋地址'], []).append(r)

    def _read_room_tags(self):
        for row in TableReader(self._db_wb['房屋标签']):
            if row['出租房']:
                self._room_tags[row['简化地址']] = '出租房'
            if row['空关房']:
                self._room_tags[row['简化地址']] = '空关房'

    def export(self, simple_address=True, merge_address=False, first_resident_address_only=True):
        tr_address = TableReader(self._db_wb['所有房屋地址'])
        col_addr = ['详细地址', '简化地址'][simple_address and 1 or 0]

        last_unit_addr = None
        writer = None
        for row in tr_address:
            cur_unit_addr = row['单元地址']
            if cur_unit_addr != last_unit_addr:
                if writer:
                    writer.save(merge_address)
                writer = DocumentWriter(self._template_path,
                                        os.path.join(
                                            self._output_dir, row['小区'], cur_unit_addr + '.xlsx'),
                                        has_outline=self._has_outline,
                                        has_room_tag=self._has_room_tag,
                                        separate_room_tag=self._separate_room_tag)
                writer.set_unit_addr(cur_unit_addr)

            writer.add_room(
                row[col_addr], self._room_tags.get(row['简化地址'], ''))
            for r in self._residents.get(row['简化地址'], []):
                writer.add_resident(row[col_addr], r)

            last_unit_addr = cur_unit_addr

        if writer:
            writer.save(merge_address, first_resident_address_only)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('template_path')
    parser.add_argument(
        'community_xlsx', help='path to the community data excel file')
    parser.add_argument('output_directory',
                        help='path to the output directory')
    parser.add_argument('--merge-address', action='store_true',
                        default=False, help='merge cells with the same address')
    parser.add_argument('--comment-tags', nargs='+',
                        help='tags to write as comments. A tag is a regex whose first group if any or the whole matched string will be written as the comment')
    parser.add_argument('--exclude-tags', nargs='+',
                        help='tags to exlude from the output')
    parser.add_argument('--simple-address', action='store_true', default=True)
    parser.add_argument('--no-room-tag', action='store_true', default=False)
    parser.add_argument('--no-outline', action='store_true', default=False)
    parser.add_argument('--no-separate-room-tag', action='store_true', default=False, help='append the tag to the comment')
    parser.add_argument('--first-resident-address-only', action='store_true', default=False)
    args = parser.parse_args()
    Exporter(args.template_path, args.community_xlsx, args.output_directory, 
             has_outline=not args.no_outline, 
             has_room_tag=not args.no_room_tag,
             separate_room_tag=not args.no_separate_room_tag,
             comment_tags=args.comment_tags, 
             exclude_tags=args.exclude_tags).export(args.simple_address, args.merge_address, args.first_resident_address_only)


if __name__ == '__main__':
    main()
