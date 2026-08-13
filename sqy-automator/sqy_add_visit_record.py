#!/usr/bin/env python3
"""Add a visit record on sqy (https://jczl.sh.cegn.cn/web/#/login).

Step 1: Open the login page in a visible browser and wait for the user to
scan the QR code with their phone. Once scanned, the browser is redirected
away from the login page and the script proceeds to add the visit record.
"""

import argparse
import datetime
import json
import sys
import time
from itertools import zip_longest

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

LOGIN_URL = "https://jczl.sh.cegn.cn/web/#/login"
LOGIN_HASH = "#/login"
LOGIN_TIMEOUT_SECONDS = 5 * 60  # how long to wait for the user to scan

VISIT_RECORD_URL = "https://jczl.sh.cegn.cn/web/#/jczlplatform/rcgz/syjdzf"
DATA_API = "/sqy-admin/api/sqReceptionVisit"  # data request that signals page load
PAGE_LOAD_TIMEOUT_MS = 30_000  # how long to wait for that request to finish

NEW_RECORD_BUTTON = "button.el-button--primary.filter-item:has-text('新增')"
VISIT_DIALOG = "div.el-dialog__wrapper.sqVisitDialog"

# 走访对象（选择居民）对话框。teleport 到 <body>，aria-label 唯一，正好区别于
# 只是包含“选择居民”文字的访客对话框。
RESIDENT_DIALOG = ".el-dialog[aria-label='选择居民']"
RESIDENT_SEARCH_INPUT = "input[placeholder='输入用户姓名检索']"
RESIDENT_SEARCH_BUTTON = "button:has-text('搜索')"
RESIDENT_SEARCH_API = "/sqy-admin/api/sqReceptionVisit/queryPersonList"
RESIDENT_SEARCH_RETRIES = 3        # 服务器不稳定，搜索请求失败时的重试次数
RESIDENT_SEARCH_RETRY_INTERVAL = 1  # 每次重试之间的间隔（秒）

SUBMIT_API = "/sqy-admin/api/bfForm/saveVisit"  # 提交走访记录的请求
CONFIRM_BUTTON = "div.el-dialog__footer button:has-text('确认')"  # 对话框底部“确认”按钮
BASE_FIELDS = {
    "走访对象", "居住地址", "方式", "走访时间", "参与人员", "走访详情", "服务标签",
}


class RecordSkipped(Exception):
    """走访记录因无法填写（如找不到走访对象）而被跳过。"""


# 示例记录；运行时可按需修改。参与人员用空格分隔的姓名列表。
EXAMPLE_RECORD = {
    "走访对象": "黄曼君",  # 需能在“选择居民”检索框找到的姓名
    "居住地址": "南大路18弄5号302",  # 必填；用于匹配“地址”列（匹配前去掉空白与末尾“室”）
    "方式": "走访",  # 需是 方式 下拉框里存在的选项
    "走访时间": "2026-08-12 14:30:00",  # 需与日期时间选择器的格式一致
    "参与人员": "李凯 朱晓庆",  # 空格分隔；姓名需在下拉框里存在
    "走访详情": "上门了解老人近期生活状况。",
    "服务标签": [
        { "tag": "困难老年人探访关爱",
          "探访关爱服务内容": "居家安全服务",
          "服务时长(小时)": 0 },
    ]
}

SERVICE_TAG_FORM_STRUCTURES = {
    "困难老年人探访关爱": {
        "form_title": "请填写困难老年人探访关爱服务信息",
        "探访关爱服务内容" : "checkbox",
        "服务时长(小时)": "input"
    }
}


def current_url(page):
    """Return the live URL from the page.

    Reads window.location.href instead of page.url because Playwright's
    page.url is not reliably refreshed on hash-only SPA navigation
    (#/login -> #/home/new), while the DOM always reflects the true URL.
    """
    return page.evaluate("window.location.href")


def wait_for_login_redirect(page, timeout_seconds=LOGIN_TIMEOUT_SECONDS):
    """Block until the page URL no longer points at the login page."""
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        url = current_url(page)
        if LOGIN_HASH not in url:
            return url
        time.sleep(1)
    raise TimeoutError(f"QR login not completed within {timeout_seconds}s")


def wait_for_visit_record_data(page, timeout_ms=PAGE_LOAD_TIMEOUT_MS):
    """Navigate to the 走访登记平台 page and wait for its data to load.

    The SPA is hash-routed, so changing the route fires an XHR for
    /sqy-admin/api/sqReceptionVisit. We start listening for that response
    before navigating, then block until it completes.
    """
    with page.expect_response(
        lambda r: DATA_API in r.url, timeout=timeout_ms
    ) as resp_info:
        page.goto(VISIT_RECORD_URL, wait_until="domcontentloaded")
    return resp_info.value


def open_new_record_dialog(page, timeout_ms=PAGE_LOAD_TIMEOUT_MS):
    """Click the 新增 button and wait for the visit-record dialog to show up.

    Element UI keeps the dialog wrapper in the DOM but hidden (display:none)
    until opened, so we wait for it to become visible rather than just exist.
    Returns the visible dialog locator, for fill_visit_record_form and
    submit_visit_record to operate on.
    """
    page.locator(NEW_RECORD_BUTTON).click()
    page.wait_for_selector(VISIT_DIALOG, state="visible", timeout=timeout_ms)
    print("新增对话框已打开")
    return page.locator(VISIT_DIALOG)


def _form_item(dialog, label):
    """Return the .el-form-item whose label text contains `label` (unique per field).

    1.15 的 locator() 不接受 has_text 参数，改用 :has-text() 伪类（同为子串匹配）。
    """
    return dialog.locator(f".el-form-item:has-text('{label}')")


def _open_dropdown_and_get(page, form_item, timeout_ms=5000):
    """Open an el-select and return its visible dropdown.

    Element UI copies the dropdown from a template and teleports it to <body>
    (class .el-select-dropdown.el-popper, display not none when open), so locate
    it at page level rather than inside the select.
    """
    form_item.locator(".el-select").click()
    # 1.15 没有 locator.wait_for，用 page.wait_for_selector 等待下拉出现
    page.wait_for_selector(
        ".el-select-dropdown.el-popper:visible", state="visible", timeout=timeout_ms
    )
    # 模板里可能还有一个 display:none 的下拉副本，挑可见的那个返回
    dropdowns = page.locator(".el-select-dropdown.el-popper")
    for i in range(dropdowns.count()):
        if dropdowns.nth(i).is_visible():
            return dropdowns.nth(i)
    raise TimeoutError("下拉列表未出现")


def set_visit_type(page, dialog, value):
    """Fill the single-select 方式 field; single-select closes automatically."""
    dropdown = _open_dropdown_and_get(page, _form_item(dialog, "方式"))
    dropdown.locator(f".el-select-dropdown__item:has-text('{value}')").click()


def set_visit_time(page, dialog, value):
    """Fill the 走访时间 datetime input and confirm with the picker's 确定 button.

    Element UI's datetime picker does not commit or close on Enter; the panel
    stays open until its footer 确定 button is clicked. That button lives at
    page level (the picker is teleported to <body>), so `page` is required.
    """
    inp = _form_item(dialog, "走访时间").locator("input.el-input__inner")
    inp.fill(value)
    inp.press("Enter")
    page.locator(
        ".el-picker-panel__footer .el-picker-panel__link-btn:has-text('确定')"
    ).click()


def set_join_users(page, dialog, users):
    """Set the multi-select 参与人员 to exactly the given names.

    `users` is a whitespace-separated name string (e.g. "李凯 朱晓庆"), or a
    list. The dialog usually starts with the current user shown as a tag; any
    currently-selected name not in `users` is removed via its close icon, then
    the listed names are picked from the dropdown.
    """
    target = set(users.split()) if isinstance(users, str) else set(users)
    form_item = _form_item(dialog, "参与人员")

    # Remove current tags whose name is not in the target list.
    while True:
        tag_to_remove = None
        tags = form_item.locator(".el-tag")
        for i in range(tags.count()):
            name = tags.nth(i).locator(".el-select__tags-text").inner_text().strip()
            if name not in target:
                tag_to_remove = tags.nth(i)
                break
        if tag_to_remove is None:
            break
        tag_to_remove.locator(".el-tag__close").click()

    # Pick the listed names from the dropdown (multi-select stays open on click).
    dropdown = _open_dropdown_and_get(page, form_item)
    items = dropdown.locator(".el-select-dropdown__item")
    for i in range(items.count()):
        item = items.nth(i)
        name = item.inner_text().strip()
        selected = "selected" in (item.get_attribute("class") or "").split()
        if name in target and not selected:
            item.click()

    # Close the dropdown with Escape
    page.keyboard.press("Escape")

    picked = set(form_item.locator(".el-select__tags-text").all_text_contents())
    missing = target - picked
    if missing:
        print(f"警告: 参与人员下拉中未找到/未选中: {missing}", file=sys.stderr)


def set_visit_content(dialog, value):
    """Fill the 走访详情 textarea."""
    _form_item(dialog, "走访详情").locator("textarea").fill(value)


def _set_checkbox_group(container, values, warn_label):
    """把 `container` 内的 checkbox 组设为恰好选中 `values` 中的选项。

    `values` 是选项名（空格分隔的字符串）或选项名列表。对每个 `.el-checkbox`：
    其 `.el-checkbox__label` 文本在目标集合里就选中（已选中则跳过），不在集合里
    但已选中就取消。请求的选项找不到对应 checkbox 时打印警告。返回实际存在的
    选项名集合（供调用方判断缺失项）。
    """
    target = set(values.split()) if isinstance(values, str) else set(values)
    boxes = container.locator(".el-checkbox")
    found = set()
    for i in range(boxes.count()):
        box = boxes.nth(i)
        label = box.locator(".el-checkbox__label").inner_text().strip()
        checked = "is-checked" in (box.get_attribute("class") or "").split()
        if label in target:
            found.add(label)
            if not checked:
                box.click()
        elif checked:
            box.click()
    missing = target - found
    if missing:
        print(f"警告: {warn_label} 中未找到选项: {sorted(missing)}", file=sys.stderr)
    return found


def _set_component_input(component, value):
    """填写组件内的输入框（el-input / el-input-number 共用 input.el-input__inner）。"""
    component.locator("input.el-input__inner").fill(str(value))


def set_service_tags(page, dialog, tags_data, timeout_ms=PAGE_LOAD_TIMEOUT_MS):
    """填写新增接待走访对话框里的 服务标签 及其服务表单。

    `tags_data` 即记录里的“服务标签”列表，每项是一个 dict：`tag` 是服务标签名，
    其余键是服务表单的字段。先把 tag-section 里的 checkbox 调整为恰好选中数据中
    的标签（数据里给出但页面没有该 checkbox 的打印警告），再对每个出现的标签按其
    form-section（用 SERVICE_TAG_FORM_STRUCTURES[tag]["form_title"] 定位）逐字段
    按声明的类型填写。form-section 只对已勾选的标签渲染，因此必须先调整 tag
    checkbox。
    """
    service_item = _form_item(dialog, "服务标签")
    tag_section = service_item.locator(".tag-section")
    tags = [e["tag"] for e in tags_data if isinstance(e, dict) and "tag" in e]
    found = _set_checkbox_group(tag_section, tags, "服务标签")

    for entry in tags_data:
        if not isinstance(entry, dict) or "tag" not in entry:
            print(f"警告: 服务标签项缺少 tag 字段，已跳过: {entry}", file=sys.stderr)
            continue
        tag = entry["tag"]
        if tag not in found:
            continue  # 缺失 checkbox 的警告已在 _set_checkbox_group 打印
        structure = SERVICE_TAG_FORM_STRUCTURES.get(tag)
        if structure is None:
            print(
                f"警告: 服务标签 {tag} 未在 SERVICE_TAG_FORM_STRUCTURES 中定义，跳过",
                file=sys.stderr,
            )
            continue
        form_title = structure["form_title"]
        section_selector = f".form-section:has-text('{form_title}')"
        try:
            page.wait_for_selector(section_selector, state="visible", timeout=timeout_ms)
        except TimeoutError:
            print(f"警告: 服务标签 {tag} 的 form-section 未出现，跳过", file=sys.stderr)
            continue
        section = service_item.locator(section_selector)
        for field_name, value in entry.items():
            if field_name == "tag":
                continue
            field_type = structure.get(field_name)
            if field_type is None:
                print(
                    f"警告: 服务标签 {tag} 的字段 {field_name} 未在结构中定义类型，跳过",
                    file=sys.stderr,
                )
                continue
            component = section.locator(f".component:has-text('{field_name}')")
            if field_type == "checkbox":
                _set_checkbox_group(component, value, field_name)
            elif field_type == "input":
                _set_component_input(component, value)
            else:
                print(
                    f"警告: 服务标签 {tag} 字段 {field_name} 的未知类型 {field_type}，跳过",
                    file=sys.stderr,
                )


def normalize_address(s):
    """归一化地址用于比较：去掉全部空白，再去掉末尾单个“室”。

    录入的“居住地址”可能有空白、可能以“室”结尾，而搜索结果的“地址”列
    不带“室”（防御性地同样处理）。
    """
    s = "".join(s.split())
    if s.endswith("室"):
        s = s[:-1]
    return s


def set_visit_target(page, dialog, name, address, timeout_ms=PAGE_LOAD_TIMEOUT_MS):
    """在“走访对象”表单项里选择姓名为 `name` 的居民。

    `address` 必填（记录里的“居住地址”字段，预期非空），用于在重名结果里精确
    匹配“地址”列；匹配前对两侧地址做归一化（去掉全部空白与末尾单个“室”）。
    点表单项内的“选择居民”按钮会打开第二个
    Element UI 对话框（teleport 到 <body>，以 aria-label 区分）。输入姓名并点“搜索”，
    等 /queryPersonList 响应返回且 JSON status==200（结果数据加密，不解析内容）。
    查无此人（结果表处于“暂无数据”或匹配不到行）时：关闭对话框、打印错误，并抛
    RecordSkipped 跳过当前记录。
    """
    _form_item(dialog, "走访对象").locator("button:has-text('选择居民')").click()
    page.wait_for_selector(RESIDENT_DIALOG, state="visible", timeout=timeout_ms)
    resident_dialog = page.locator(RESIDENT_DIALOG)

    def _skip(reason):
        resident_dialog.locator(".el-dialog__headerbtn").click()
        raise RecordSkipped(reason)

    # 结果数据加密，只能判断请求是否成功：以 queryPersonList 响应返回且其 JSON 的
    # status 字段为 200 作为搜索完成信号（不解析加密的 data）。服务器不稳定时
    # 请求可能超时或 status != 200，重试 RESIDENT_SEARCH_RETRIES 次。
    resident_dialog.locator(RESIDENT_SEARCH_INPUT).fill(name)
    last_reason = None
    for attempt in range(1, RESIDENT_SEARCH_RETRIES + 1):
        ok = False
        try:
            with page.expect_response(
                lambda r: RESIDENT_SEARCH_API in r.url, timeout=timeout_ms
            ) as resp_info:
                resident_dialog.locator(RESIDENT_SEARCH_BUTTON).click()
            if resp_info.value.json().get("status") == 200:
                ok = True
            else:
                last_reason = "status != 200"
        except TimeoutError:
            last_reason = f"{timeout_ms}ms 内未收到响应"
        if ok:
            break
        if attempt < RESIDENT_SEARCH_RETRIES:
            print(
                f"走访对象搜索第 {attempt} 次失败（{last_reason}），"
                f"{RESIDENT_SEARCH_RETRY_INTERVAL}s 后重试 ...",
                file=sys.stderr,
            )
            time.sleep(RESIDENT_SEARCH_RETRY_INTERVAL)
    else:
        _skip(
            f"走访对象搜索请求 {RESIDENT_SEARCH_RETRIES} 次均失败"
            f"（{last_reason}），跳过该条记录"
        )

    # 轮询等数据行渲染，按姓名（以及可选的地址）匹配目标行；结果表持续处于
    # “暂无数据”空状态（tbody 无数据行且空状态文本连续可见）时判为查无此人。
    rows = resident_dialog.locator("tbody tr")
    empty_text = resident_dialog.locator(".el-table__empty-text")
    target_row = None
    deadline = time.time() + timeout_ms / 1000
    empty_ticks = 0
    while time.time() < deadline:
        for i in range(rows.count()):
            cells = rows.nth(i).locator("td")
            if cells.count() < 2:
                continue  # 防御：仅一列的空态行
            if cells.nth(0).inner_text().strip() != name:
                continue
            if normalize_address(cells.nth(2).inner_text()) != normalize_address(address):
                continue
            target_row = rows.nth(i)
            break
        if target_row is not None:
            break
        if rows.count() == 0 and empty_text.count() > 0 and empty_text.first.is_visible():
            empty_ticks += 1
            if empty_ticks >= 3:  # 连续约 0.6s 仍为空，判定查无此人
                break
        else:
            empty_ticks = 0
        time.sleep(0.2)
    if target_row is None:
        _skip(f"走访对象 {name} 未找到（暂无数据或没有匹配行），跳过该条记录")

    # 点“操作”列（最后一列）里的“选择”按钮；找不到时退而点该行最后一个按钮。
    action_btn = target_row.locator("td:last-child button:has-text('选择')")
    if action_btn.count() == 0:
        action_btn = target_row.locator("td:last-child button")
    action_btn.click()

    # 等“选择居民”对话框关闭，再读回已选姓名打印确认。
    page.wait_for_selector(RESIDENT_DIALOG, state="hidden", timeout=timeout_ms)
    print(f"已设置走访对象: {_form_item(dialog, '走访对象').inner_text().strip()}")


def fill_visit_record_form(page, dialog, record, timeout_ms=PAGE_LOAD_TIMEOUT_MS):
    """Fill the open 新增接待走访 dialog from `record`.

    `dialog` is the locator returned by open_new_record_dialog. Supported keys:
    走访对象, 居住地址, 方式, 走访时间, 参与人员, 走访详情, 服务标签.
    走访对象 must come first: if the resident can't be found, RecordSkipped is
    raised and the remaining fields are left untouched (the record is skipped).
    走访对象 要求记录里同时提供居住地址（用于在结果表里精确匹配“地址”列）。
    服务标签 是列表，每项含 tag 与对应服务字段；tag 的 checkbox 及每个字段的
    填写按 SERVICE_TAG_FORM_STRUCTURES 定义的类型进行。
    Keys absent from `record` are left untouched.
    """
    if "走访对象" in record:
        set_visit_target(page, dialog, record["走访对象"], record["居住地址"])
    if "方式" in record:
        set_visit_type(page, dialog, record["方式"])
    if "走访时间" in record:
        set_visit_time(page, dialog, record["走访时间"])
    if "参与人员" in record:
        set_join_users(page, dialog, record["参与人员"])
    if "走访详情" in record:
        set_visit_content(dialog, record["走访详情"])
    if "服务标签" in record:
        set_service_tags(page, dialog, record["服务标签"])


def parse_service_tags(value):
    """把 服务标签 单元格解析成记录里的列表形式。

    单元格是 JSON 字符串（结构同 EXAMPLE_RECORD 的 服务标签：dict 列表，每个含 tag
    与服务表单字段）。JSON 解析失败或非 JSON 字符串时返回空列表。非字符串原样返回。
    """
    if isinstance(value, str):
        text = value.strip()
        if not (text.startswith("[") or text.startswith("{")):
            return []
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError as exc:
            print(f"警告: 服务标签 JSON 解析失败（{exc}），返回空列表",
                  file=sys.stderr)
            return []
        if isinstance(parsed, dict):
            parsed = [parsed]
        return parsed
    return value


def parse_record(raw):
    """把一行原始单元格值整理成 fill_visit_record_form 能用的记录 dict。

    去掉 None / 全空白的值；走访时间 若是 Excel 日期对象则格式化为
    %Y-%m-%d %H:%M:%S 以匹配日期时间选择器的格式；服务标签 经 parse_service_tags
    解析成列表。
    """
    record = {}
    for key, value in raw.items():
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        record[key] = value
    if "走访时间" in record and isinstance(record["走访时间"], datetime.datetime):
        record["走访时间"] = record["走访时间"].strftime("%Y-%m-%d %H:%M:%S")
    if "服务标签" in record:
        record["服务标签"] = parse_service_tags(record["服务标签"])
    return record


def read_visit_records(path, months):
    """从 xlsx 读取走访记录，按月份（表）与行序返回记录 dict 列表。

    每个工作表名为 `X月`。第一行为表头，列名即记录字段（走访对象、居住地址、方式、
    走访时间、参与人员、走访详情、服务标签）；其后每行是一条记录。指定的月份工作表
    不存在时打印错误并以非零码退出；表头里非示例字段的列打印警告并被忽略。
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"错误: 找不到文件 {path}", file=sys.stderr)
        sys.exit(1)
    records = []
    for month in months:
        title = f"{month}月"
        if title not in wb.sheetnames:
            print(f"错误: 工作表 {title} 不存在，可用工作表: {wb.sheetnames}",
                  file=sys.stderr)
            sys.exit(1)
        ws = wb[title]
        rows = ws.iter_rows(values_only=True)
        header = [(str(c).strip() if c is not None else "") for c in next(rows, ())]
        for col in header:
            if col and col not in BASE_FIELDS:
                print(f"警告: 工作表 {title} 的列 {col} 不是示例记录的字段，将被忽略",
                      file=sys.stderr)
        for row in rows:
            raw = {}
            for col, value in zip_longest(header, row):
                if not col:
                    continue
                raw[col] = value
            if raw:
                records.append(parse_record(raw))
    return records


def _valid_month(text):
    try:
        m = int(text)
    except ValueError:
        print(f"错误: 月份参数不是数字: {text}", file=sys.stderr)
        sys.exit(1)
    if not 1 <= m <= 12:
        print(f"错误: 月份超出范围 1-12: {text}", file=sys.stderr)
        sys.exit(1)
    return m


def parse_months(raw_values):
    """把 --months 参数展开为排序去重的月份数字列表。

    每项可以是单个月份数字，或 `from,to` 的闭区间（含 to）。省略时默认当前月份。
    """
    if not raw_values:
        return [datetime.date.today().month]
    months = set()
    for item in raw_values:
        parts = item.split(",")
        if len(parts) == 1:
            months.add(_valid_month(item))
        elif len(parts) == 2:
            lo, hi = _valid_month(parts[0]), _valid_month(parts[1])
            if lo > hi:
                print(f"错误: 月份区间 {item} 的起始月大于结束月", file=sys.stderr)
                sys.exit(1)
            months.update(range(lo, hi + 1))
        else:
            print(f"错误: 无法识别的月份参数: {item}", file=sys.stderr)
            sys.exit(1)
    return sorted(months)


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="在 sqy 上新增接待走访记录；从 xlsx 读取记录并逐条填写、提交。"
    )
    parser.add_argument(
        "-i", "--input",
        metavar="XLSX",
        help="走访记录 xlsx 文件路径；省略则使用内置示例记录 EXAMPLE_RECORD",
    )
    parser.add_argument(
        "--months",
        nargs="*",
        metavar="M",
        help="要读取的工作表月份（对应 X月 工作表），可为单个数字或 from,to 闭区间；"
             "省略则默认当前月份",
    )
    parser.add_argument(
        "--confirm",
        action="store_true",
        help="每填完一条记录后暂停，按回车确认后再提交并继续下一条",
    )
    args = parser.parse_args(argv)
    args.months = parse_months(args.months)
    return args


def close_visit_dialog(page, timeout_ms=5000):
    """关闭“新增接待走访”对话框并等它隐藏（尽力而为）。

    优先点右上角关闭按钮；找不到时按 Escape（Element UI 默认支持）。用于某条记录
    跳过/提交失败后，保证下一条记录能重新打开对话框。关闭失败只打印警告，不中断
    后续记录。
    """
    dialog = page.locator(VISIT_DIALOG)
    if dialog.is_visible():
        close_btn = dialog.locator(".el-dialog__headerbtn")
        if close_btn.count() > 0:
            close_btn.click()
        else:
            page.keyboard.press("Escape")
    try:
        page.wait_for_selector(VISIT_DIALOG, state="hidden", timeout=timeout_ms)
    except TimeoutError:
        print("警告: 对话框未能自动关闭", file=sys.stderr)


def submit_visit_record(page, dialog, record, timeout_ms=PAGE_LOAD_TIMEOUT_MS):
    """点对话框底部“确认”按钮提交当前走访记录，返回是否成功。

    `dialog` 是 open_new_record_dialog 返回的 locator。提交后等待
    /sqy-admin/api/bfForm/saveVisit 响应；其 JSON 的 status 字段为 200 视为成功
    （判读方式与搜索请求一致），否则打印错误并返回 False。提交成功后对话框通常自动
    关闭；若未关闭则点右上角 X 关掉，保证下一条能重开。`record` 仅用于错误提示里
    定位是哪条记录。
    """
    confirm_btn = dialog.locator(CONFIRM_BUTTON)
    if confirm_btn.count() == 0:
        confirm_btn = dialog.locator("button:has-text('确认')")
    if confirm_btn.count() == 0:
        raise TimeoutError("未找到“确认”按钮")

    with page.expect_response(
        lambda r: SUBMIT_API in r.url, timeout=timeout_ms
    ) as resp_info:
        confirm_btn.click()
    resp = resp_info.value
    try:
        data = resp.json()
        status = data.get("status") if isinstance(data, dict) else None
    except ValueError:
        status = None
    if status == 200:
        print("提交成功")
        ok = True
    else:
        print(f"错误: 保存请求失败（status={status}），记录: {record.get('走访对象')}",
              file=sys.stderr)
        ok = False

    # 保存成功后对话框通常自动关闭；没关闭就关掉它。
    try:
        page.wait_for_selector(VISIT_DIALOG, state="hidden", timeout=5000)
    except TimeoutError:
        close_visit_dialog(page)
    return ok


def main():
    args = parse_args()
    if args.input:
        records = read_visit_records(args.input, args.months)
    else:
        records = [EXAMPLE_RECORD]
    print(f"共读取 {len(records)} 条记录")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.goto(LOGIN_URL)

        print("请在浏览器中扫描二维码完成登录 ...")
        try:
            final_url = wait_for_login_redirect(page)
        except TimeoutError as exc:
            print(exc, file=sys.stderr)
            browser.close()
            sys.exit(1)

        print(f"登录成功，已跳转到: {final_url}")

        resp = wait_for_visit_record_data(page)
        print(f"已进入平台，{DATA_API} 返回状态码: {resp.status}")

        success_count = 0
        failed_count = 0
        for i, record in enumerate(records, 1):
            dialog = open_new_record_dialog(page)
            try:
                fill_visit_record_form(page, dialog, record)
                if args.confirm:
                    input(
                        f"已填写第 {i}/{len(records)} 条记录，"
                        f"请检查表单，按回车提交并继续..."
                    )
                if submit_visit_record(page, dialog, record):
                    success_count += 1
                else:
                    failed_count += 1
            except RecordSkipped as exc:
                failed_count += 1
                print(f"错误: {exc}", file=sys.stderr)
                close_visit_dialog(page)
            except TimeoutError as exc:
                failed_count += 1
                print(f"错误: {exc}", file=sys.stderr)
                close_visit_dialog(page)

        print(f"共 {len(records)} 条记录：成功 {success_count} 条，失败/跳过 {failed_count} 条")
        print("浏览器保持打开供检查，按回车退出...")
        input()
        browser.close()


if __name__ == "__main__":
    main()
