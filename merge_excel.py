#! /usr/bin/env python3

import glob
import argparse
import openpyxl


def cell_address(s):
    return tuple(int(e) for e in s.split(','))


def copy_range(src_ws: openpyxl.worksheet.worksheet.Worksheet,
               src_row, src_col_start, src_col_end,
               dest_ws: openpyxl.worksheet.worksheet.Worksheet, 
               dest_row, dest_col):
    '''
    copy the source range to the destination range
    returns True if any non empty cell is copied
    '''
    col = src_col_start
    copied = False
    while col <= src_col_end:
        c = src_ws.cell(src_row, col)
        if c.value:
            dest_ws.cell(dest_row, dest_col).value = c.value
            copied = True
        col += 1
        dest_col += 1
    return copied


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('source', nargs='+',
                        help='source workbook path, simple unix shell style \
                              pattern can be used')
    parser.add_argument('--top-left', metavar='ROW,COL', type=cell_address,
                        help='the top left cell of the range in \
                              each worksheet to merge, i.e. row,col')
    parser.add_argument('--header', action='store_true', default=False,
                        help='if the first row is a header row')
    parser.add_argument('-w', type=int, dest='width',
                        help='width of the range to merge')
    parser.add_argument('-o', required=True, dest='output',
                        help='output workbook path')
    args = parser.parse_args()

    output = openpyxl.Workbook()
    out_ws = output.worksheets[0]
    first = True
    range_width = args.width
    out_row = 1
    for p in args.source:
        for f in glob.glob(p):
            ws = openpyxl.load_workbook(f).worksheets[0]
            row, col = args.top_left
            if not range_width:
                range_width = ws.max_column - col + 1
            if args.header:
                if first:
                    first = False
                else:
                    row += 1
            while row <= ws.max_row and \
                    copy_range(ws, row, col, col + range_width - 1,
                               out_ws, out_row, 1):
                row += 1
                out_row += 1
    output.save(args.output)
